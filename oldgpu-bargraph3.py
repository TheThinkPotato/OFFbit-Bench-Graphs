from turtle import *
from pathlib import Path
import openpyxl
#from threading import thread

type_number = 0

types = ["Cloudgate","Conter Strike Source","Cinebench 11.5"]
type = types[type_number]


color1 = "#c0504d"
color2 = "#4f81bd"
color_h1 = "#e29300"
color_h2 = "#ba00d7"

setup(1200,600)
speed("fastest")
title("CPU Benchmarks")
tracer(10)

hideturtle()

start_x = -300
start_y = -20

cpu_names_gpu = []
threeDmark = []
CS_source = []
cinebench_gpu = []

bar_width = 25

# Get data from spreadsheet

def read_spread_sheet_gpu():
    global cpu_names_gpu
    global threeDmark
    global CS_source
    global cinebench_gpu

    xlsx_file = Path('gpu-Bench Graph Old Card.xlsx')
    wb_obj = openpyxl.load_workbook(xlsx_file) 

    # Read the active sheet:
    sheet = wb_obj.active

    index = 0
    while index != 6 :
        cpu_names_gpu.append((sheet.cell(row = 2+index , column= 2).value))
        threeDmark.append(sheet.cell(row = 2+index , column= 3).value)     
        CS_source.append(sheet.cell(row = 2+index , column= 4).value)     
        cinebench_gpu.append(sheet.cell(row = 2+index , column= 5).value )    
        index += 1

# Title
def draw_title(text = 'Insert Title'):
    global start_x
    global start_y
    pu()
    style = ('Arial', 20, 'bold')
    goto(0,270+start_y)
    pd()
    write(text, font=style, align='center')
    pu()

# Draw Axis Lines
def xy_line():
    global start_x
    global start_y
    color('gray')
    width(2)
    pu()
    goto(start_x,start_y + 260)
    pd()
    rt(90)
    fd(505)
    lt(90)
    #fd(800)
    pu()
    width(1)
    color('black')

# Draw legend
def draw_legend(icons = 2, name_1 = "Single Core", name_2 = "Multi Core", offset_x = 860, offset_y = 20):
    style = ('Arial', 14, 'italic')
    global start_x
    global start_y
    offset_y += start_y
    legend_x = start_x + offset_x
    pu()
    goto(legend_x, offset_y)
 
    if icons == 2:
        fillcolor(color2)
        begin_fill()
        pd()
        for lines in range(4):
            fd(10)
            lt(90)
        end_fill()
        pu()
        goto(legend_x-5, offset_y-7)
        write(name_1, font=style, align='right')

        goto(legend_x, offset_y - 30)
        fillcolor(color1)
        begin_fill()
        pd()
        for lines in range(4):
            fd(10)
            lt(90)
        end_fill()
        pu()
        goto(legend_x-5, offset_y-30-7)
        write(name_2, font=style, align='right')

    elif icons == 1:
        fillcolor(color2)
        begin_fill()
        pd()
        for lines in range(4):
            fd(10)
            lt(90)
        end_fill()
        pu()
        goto(legend_x-5, offset_y-7)
        write(name_1, font=style, align='right')

# Draw bar graphs
def draw_bar_data(bar_data):
    global start_x
    global start_y
    global color1
    global color2
    global color_h1
    global color_h2

    first_color = 2
    
    color('black')
    current_colour = color2
    current_gap = 25
    group_gap = 10
    normal_gap = 25
    y_offset = 250 + start_y
    pu()
    goto(start_x,y_offset)
    pd()
    scaled_data = scale_data(bar_data)
    
    for bar in scaled_data:
        if current_colour == color2:
            current_colour = color1
        else:
            current_colour = color2

        if first_color == 2:
            current_colour = color_h1
            first_color -= 1
        elif first_color == 1:
            current_colour = color_h2
            first_color -= 1
        
        fillcolor(current_colour)
        begin_fill()
        fd(bar)
        rt(90)
        fd(bar_width)
        rt(90)
        fd(bar)
        rt(90)
        fd(bar_width)
        rt(90)
        pu()
        end_fill()

        if current_gap == 25:
            current_gap = group_gap
        else:
            current_gap = normal_gap
        y_offset -= bar_width + current_gap
        goto(start_x,y_offset)
        pd()
    pu()

# Draw bar graphs
def draw_bar_data_single(bar_data):
    global start_x
    global start_y
    global color1
    global color2
    global color_h1
    global color_h2
    fist_color = 1
    color('black')
    current_colour = color2

    current_gap = 60
    y_offset = 230 + start_y

    pu()
    goto(start_x,y_offset)
    pd()
    scaled_data = scale_data(bar_data)
    
    for bar in scaled_data:
        if fist_color == 1:
            current_colour = color_h2
            fist_color -= 1
        else:
            current_colour = color2    
        fillcolor(current_colour)
        
        begin_fill()
        fd(bar)
        rt(90)
        fd(bar_width)
        rt(90)
        fd(bar)
        rt(90)
        fd(bar_width)
        rt(90)
        pu()
        end_fill()

        y_offset -= bar_width + current_gap
        goto(start_x,y_offset)
        pd()
    pu()

# Draw bar graph text values
def write_data_text(bar_data):
    global start_x
    global start_y
    current_gap = 25
    group_gap = 10
    normal_gap = 25
    y_offset = 225 + start_y
    style = ('Arial', 16, 'italic')
    color('white')

    scaled_data = scale_data(bar_data)
    index = 0

    for bar_amount in scaled_data:
        goto(start_x+bar_amount-5,y_offset)
        if bar_amount != 0:
            write(bar_data[index], font=style, align='right')
        index += 1
        if current_gap == 25:
            current_gap = group_gap
        else:
            current_gap = normal_gap
        y_offset -= bar_width + current_gap

# Draw bar graph text values
def write_data_text_single(bar_data):
    global start_x
    global start_y
    current_gap = 60

    y_offset = 205 + start_y
    style = ('Arial', 16, 'italic')
    color('white')

    scaled_data = scale_data(bar_data)
    index = 0

    for bar_amount in scaled_data:
        goto(start_x+bar_amount-5,y_offset)
        if bar_amount != 0:
            write(bar_data[index], font=style, align='right')
        index += 1
        y_offset -= bar_width + current_gap
    
# Draw Device names on y axis
def write_names_text(text_data):
    global start_x
    global start_y
    current_gap = 85
    y_offset = 205 + start_y
    style = ('Arial', 16, 'italic')
    color('black')

    for text in text_data:
        goto(start_x -10,y_offset)
        write(text, font=style, align='right')

        y_offset -= current_gap

# Scale Graphs
def scale_data(bar_data):
    scaled_data = []
    scale = 1
    max_item = 0

    try:
        max_item = max(bar_data)    
    except:
        print("scale_data function error - get max error")

    if max_item >= 9000 and max_item < 10000:
        scale = 15

    if max_item >= 5000 and max_item < 9000:
        scale = 10

    if max_item >= 1200 and max_item < 5000:
        scale = 2

    if max_item >= 800 and max_item < 1200:
        scale = 1.5

    if max_item >= 600 and max_item < 800:
        scale = 1.1

    if max_item >= 400 and max_item < 600:
        scale = .8

    if max_item >= 200 and max_item < 400:
        scale = .5
    
    if max_item >= 80 and max_item < 200:
        scale = .14
    
    if max_item >= 50 and max_item < 80:
        scale = .08

    if max_item >= 10 and max_item < 50:
        scale = .06

    if max_item <= 10:
        scale = .01

    for item in bar_data:
        scaled_data.append(item//scale)

    return scaled_data

# Main Draw to screen
def screen_draw():
    clear()
    xy_line()
    
    if type == "Cloudgate":
        write_names_text(cpu_names_gpu)
        draw_legend(1,"GPU Points")
        title_name = "GPU: 3DMark - Cloud Gate"
        title(title_name)
        bgpic("bg3dmark.png")
        draw_title(title_name[5:]) 
        
        draw_bar_data_single(threeDmark)
        write_data_text_single(threeDmark)
    
    elif type == "Conter Strike Source":
        write_names_text(cpu_names_gpu)
        draw_legend(1,"GPU Points")
        title_name = "GPU: Counter Strike Source"
        title(title_name)
        bgpic("bgCss.png")
        draw_title(title_name[5:]) 
        
        draw_bar_data_single(CS_source)
        write_data_text_single(CS_source)

    elif type == "Cinebench 11.5":
        write_names_text(cpu_names_gpu)
        draw_legend(1,"FPS")
        title_name = "GPU: Cinebench 11.5 GPU Benchmark"
        title(title_name)
        bgpic("bgcb.png")
        draw_title(title_name[5:]) 
        
        draw_bar_data_single(cinebench_gpu)
        write_data_text_single(cinebench_gpu)

    else:
        pass

# Change graph via keyboard presses
def image_change_up(key):
    global type
    global types
    global type_number

    if key == "Right":
        if type_number < len(types)-1:
            type_number = type_number +  1
        else:
            type_number = 0
    elif key == "Left":
        if type_number > 1:
            type_number = type_number -  1
        else:
            type_number = len(types)-1
    
    elif key == "Home":
        type_number = 0
    
    elif key == "End":
        type_number = len(types) -1
        
    else:
        pass

    type = types[type_number]
    screen_draw()

# Reads both cpu and gpu spreadsheets
def read_spreadsheets():
    try:
        read_spread_sheet_gpu()
    except:
        print('Cant Read GPU Spread sheet')

# gets keyboard inputs and updates screen
def get_keys():
    listen()
    onkeypress(lambda: image_change_up("Right"),"Right")
    onkeypress(lambda: image_change_up("Right"),"Up")
    #onkeypress(lambda: image_change_up("Right"),"PageUp")
    onkeypress(lambda: image_change_up("Left"),"Left")
    onkeypress(lambda: image_change_up("Left"),"Down")
    #onkeypress(lambda: image_change_up("Left"),"PageDown")
    onkeypress(lambda: image_change_up("Home"),"Home")
    onkeypress(lambda: image_change_up("End"),"End")


read_spreadsheets()
screen_draw()
get_keys()

mainloop()
done()

